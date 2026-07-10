import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

# Configuración del script
INPUT_FILE = 'TRX WU_BP.xlsx'
RANKING_FILE = os.path.join('Reportes_Individuales_CSV', 'Ranking.csv')
OUTPUT_FILE = 'Reporte_Salud_Top300_2026.xlsx'
INACTIVITY_WEEKS = 4
VARIATION_THRESHOLD = 15.0

def generar_reporte_top300():
    print("Iniciando análisis de Health Score TOP 300 2026...")
    if not os.path.exists(INPUT_FILE):
        print(f"Error: No se encontró el archivo de entrada '{INPUT_FILE}'")
        return
        
    # 1. Cargar la hoja Semáforo
    print("Cargando hoja de 'Semáforo'...")
    df_sem = pd.read_excel(INPUT_FILE, sheet_name='Semáforo')
    cols = df_sem.columns.tolist()
    row_0 = df_sem.iloc[0].tolist()
    
    # Reconstruir dataframe limpio
    df_sem_clean = df_sem.iloc[1:].copy()
    df_sem_clean.rename(columns={cols[0]: 'Director', cols[1]: 'Holder'}, inplace=True)
    df_sem_clean = df_sem_clean[df_sem_clean['Holder'].notna()]
    df_sem_clean = df_sem_clean[df_sem_clean['Holder'].astype(str).str.lower() != 'total']
    
    # Identificar semanas de 2026
    weeks_2026 = [c for c in cols if str(c).startswith('Semana') and '2026' in str(c)]
    
    # Convertir semanas a numérico y llenar NaNs con 0
    for col in weeks_2026:
        df_sem_clean[col] = pd.to_numeric(df_sem_clean[col], errors='coerce').fillna(0)
        
    # Usar todos los clientes (incluso los de 0 actividad si son Top 300)
    df_26 = df_sem_clean.copy()
    
    # Cargar y cruzar con Vendedor y Top 300 desde Ranking.csv
    print("Cargando Vendedores y Top 300 desde Ranking.csv...")
    vendedor_map = {}
    top300_holders = []
    if os.path.exists(RANKING_FILE):
        try:
            df_rank = pd.read_csv(RANKING_FILE, encoding='utf-8')
            if 'Holder' in df_rank.columns and 'Vendedor' in df_rank.columns:
                vendedor_map = dict(zip(df_rank['Holder'], df_rank['Vendedor']))
                print(f"  Mapeados {len(vendedor_map)} vendedores de Ranking.csv")
            if 'Holder' in df_rank.columns and 'Ranking' in df_rank.columns:
                top300_holders = df_rank[df_rank['Ranking'] <= 300]['Holder'].tolist()
        except Exception as e:
            print(f"  Advertencia: No se pudo leer {RANKING_FILE}: {e}")
    else:
        print(f"  Nota: No se encontró {RANKING_FILE}")
        
    if top300_holders:
        df_26 = df_26[df_26['Holder'].isin(top300_holders)].copy()
        print(f"Filtrado a {len(df_26)} clientes en el TOP 300.")
    else:
        # Si no hay Ranking.csv, sacar los Top 300 por volumen total aquí
        df_26['Total_Volume_2026'] = df_26[weeks_2026].sum(axis=1)
        df_26 = df_26.nlargest(300, 'Total_Volume_2026')
        print(f"Filtrado a {len(df_26)} clientes tomando los mayores de 2026 directamente.")
        
    # 3. Clasificación de Valor (Segmentación ABC)
    print("Realizando segmentación ABC para los Top 300...")
    if 'Total_Volume_2026' not in df_26.columns:
        df_26['Total_Volume_2026'] = df_26[weeks_2026].sum(axis=1)
    
    # Ordenar de mayor a menor volumen
    df_26 = df_26.sort_values(by='Total_Volume_2026', ascending=False).reset_index(drop=True)
    
    n_clients = len(df_26)
    n_a = int(np.round(n_clients * 0.10))
    n_b = int(np.round(n_clients * 0.20))
    
    df_26['Segmento'] = 'Clase C'
    df_26.loc[:n_a-1, 'Segmento'] = 'Clase A'
    df_26.loc[n_a:n_a+n_b-1, 'Segmento'] = 'Clase B'
    
    # Mapeo de semana a mes
    week_to_month = {col: month for col, month in zip(cols, row_0) if str(col).startswith('Semana')}
    last_week_limit_2026 = len(weeks_2026) - INACTIVITY_WEEKS
    last_4_weeks = weeks_2026[-4:]
    
    # 4. Cálculos detallados por cliente
    print("Calculando variaciones, pérdidas y Score Numérico de Salud (0-100)...")
    results = []
    
    for idx, row in df_26.iterrows():
        holder = row['Holder']
        director = row['Director'] if pd.notna(row['Director']) else 'Sin Director'
        vendedor = vendedor_map.get(holder, director)  # fallback a director si no se encuentra
        segmento = row['Segmento']
        
        active_w_26 = [w for w in weeks_2026 if row[w] > 0]
        
        if not active_w_26:
            # Cliente sin actividad en todo 2026
            last_w = 'Sin actividad'
            lost = True
            baseline_avg = 0.0
            recent_avg = 0.0
            pct_change = 0.0
            loss_absolute = 0.0
            trend_score = 0.0
            freq_score = 0.0
            recency_score = 0.0
            health_score = 0.0
        else:
            last_w = active_w_26[-1]
            w_num = int(last_w.split()[1])
            
            # Criterio de perdida
            recent_activity = sum(row[last_4_weeks]) > 0
            lost = not recent_activity
            
            first_w = active_w_26[0]
            first_w_idx = weeks_2026.index(first_w)
            
            # Valores de promedios
            if first_w_idx >= 21:
                baseline_avg = 0.0
                recent_avg = float(row[last_4_weeks].mean())
                pct_change = 0.0
            else:
                baseline_w = weeks_2026[first_w_idx:21]
                baseline_avg = float(row[baseline_w].mean())
                recent_avg = float(row[last_4_weeks].mean())
                pct_change = (recent_avg - baseline_avg) / baseline_avg * 100 if baseline_avg > 0 else 0.0
                
            # Pérdida absoluta en pesos/volumen
            loss_absolute = max(0.0, baseline_avg - recent_avg) if not lost else baseline_avg
            
            # --- CÁLCULO DEL HEALTH SCORE (0-100) ---
            # A) Tendencia Score (0-50 pts)
            if lost:
                trend_score = 0.0
            elif first_w_idx >= 21:
                trend_score = 35.0
            else:
                if pct_change >= VARIATION_THRESHOLD:
                    trend_score = 50.0
                elif -VARIATION_THRESHOLD <= pct_change < VARIATION_THRESHOLD:
                    trend_score = 35.0 + (pct_change + VARIATION_THRESHOLD) / (2.0 * VARIATION_THRESHOLD) * 10.0
                else:
                    trend_score = max(0.0, 35.0 * (1.0 + pct_change / 100.0))
                    
            # B) Frecuencia Score (0-30 pts)
            start_check_idx = max(13, first_w_idx)  # index 13 es Semana 14
            weeks_to_check = weeks_2026[start_check_idx:25]
            n_weeks_check = len(weeks_to_check)
            if n_weeks_check > 0:
                n_active_weeks = sum(row[weeks_to_check] > 0)
                freq_score = (n_active_weeks / n_weeks_check) * 30.0
            else:
                freq_score = 30.0
                
            # C) Recencia Score (0-20 pts)
            if last_w == 'Semana 25 2026':
                recency_score = 20.0
            elif last_w == 'Semana 24 2026':
                recency_score = 15.0
            elif last_w == 'Semana 23 2026':
                recency_score = 10.0
            elif last_w == 'Semana 22 2026':
                recency_score = 5.0
            else:
                recency_score = 0.0
                
            health_score = float(trend_score + freq_score + recency_score)
        
        # --- ESTADO Y PRIORIZACIÓN DE ATENCIÓN ---
        if lost:
            status = 'Perdido'
        elif first_w_idx >= 21:
            status = 'Constante'
        elif pct_change > VARIATION_THRESHOLD:
            status = 'Ha subido'
        elif pct_change < -VARIATION_THRESHOLD:
            status = 'Ha bajado'
        else:
            status = 'Constante'
            
        # Prioridad de atención y acciones comerciales sugueridas
        if status == 'Perdido':
            if segmento == 'Clase A':
                priority = 'CRÍTICA: Recuperar Cliente A'
                action = 'Entrevista de salida por el Director Comercial y oferta de reactivación VIP.'
            elif segmento == 'Clase B':
                priority = 'ALTA: Recuperar Cliente B'
                action = 'Contacto de recuperación por el Vendedor con incentivos de volumen.'
            else:
                priority = 'MEDIA: Recuperar Cliente C'
                action = 'Envío de correo/WhatsApp automatizado de reactivación.'
        elif status == 'Ha bajado':
            if segmento == 'Clase A':
                priority = '🚨 INMEDIATA: Retención Roja (Clase A)'
                action = 'Llamada inmediata del Director para evaluar causa de caída y proponer plan de retención/tarifa.'
            elif segmento == 'Clase B':
                priority = '⚠️ ALTA: Retención Naranja (Clase B)'
                action = 'Visita o contacto urgente del Vendedor asignado para verificar problemas operativos.'
            else:
                priority = 'MEDIA: Seguimiento Retención (Clase C)'
                action = 'Contacto telefónico por el Vendedor para revisar volumen.'
        elif status == 'Constante':
            if segmento == 'Clase A':
                priority = 'MEDIA: Monitoreo Preventivo A'
                action = 'Revisar consistencia operativa de forma quincenal.'
            else:
                priority = 'BAJA: Operación Normal'
                action = 'Monitoreo pasivo mensual.'
        else: # Ha subido
            if segmento == 'Clase A':
                priority = 'BAJA: Fidelización VIP'
                action = 'Enviar felicitación / Buscar expansión de cartera o referencias.'
            else:
                priority = 'BAJA: Operación Normal'
                action = 'Monitoreo pasivo mensual.'
                
        results.append({
            'Holder': holder,
            'Director': director,
            'Vendedor': vendedor,
            'Segmento': segmento,
            'Promedio Inicial': baseline_avg,
            'Promedio Reciente': recent_avg,
            'Variación': pct_change,
            'Pérdida Absoluta': loss_absolute,
            'Health Score': health_score,
            'Estado de Salud': status,
            'Prioridad de Atención': priority,
            'Acción Comercial Sugerida': action,
            'Última Semana Activo': last_w,
            'Último Mes Activo': week_to_month.get(last_w, 'Ninguno')
        })
        
    df_final = pd.DataFrame(results)
    
    # Crear listas de alertas
    df_alertas_rojas = df_final[df_final['Prioridad de Atención'].isin(['🚨 INMEDIATA: Retención Roja (Clase A)', 'CRÍTICA: Recuperar Cliente A'])].sort_values(by='Pérdida Absoluta', ascending=False)
    df_alertas_naranjas = df_final[df_final['Prioridad de Atención'].isin(['⚠️ ALTA: Retención Naranja (Clase B)', 'ALTA: Recuperar Cliente B'])].sort_values(by='Pérdida Absoluta', ascending=False)
    
    # 5. Crear libro Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="1F4E78")
    desc_font = Font(name=font_family, size=9, italic=True, color="595959")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    zebra_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Colores para las tarjetas KPI y estados
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    font_green_kpi = Font(name=font_family, size=18, bold=True, color="375623")
    font_red_kpi = Font(name=font_family, size=18, bold=True, color="C65911")
    font_blue_kpi = Font(name=font_family, size=18, bold=True, color="1F4E78")
    font_grey_kpi = Font(name=font_family, size=18, bold=True, color="595959")
    
    label_font = Font(name=font_family, size=9, bold=True, color="595959")
    normal_font = Font(name=font_family, size=10)
    bold_font = Font(name=font_family, size=10, bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # ----------------- HOJA 1: DASHBOARD DE SALUD -----------------
    ws_dash = wb.create_sheet(title="Dashboard de Salud")
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash['A1'] = "DASHBOARD DE SALUD COMERCIAL 2026 - TOP 300"
    ws_dash['A1'].font = title_font
    ws_dash['A2'] = f"Seguimiento preventivo y plan de acción de la cartera TOP 300 ({n_clients} clientes activos en 2026)"
    ws_dash['A2'].font = desc_font
    
    kpis_config = [
        {"col": 1, "label": "ALERTAS ROJAS 🚨", "formula": "=SUM(I14:J14)", "fill": fill_red, "font": font_red_kpi},
        {"col": 2, "label": "ALERTAS NARANJAS ⚠️", "formula": "=SUM(I15:J15)", "fill": fill_red, "font": font_red_kpi},
        {"col": 3, "label": "ESTABLES ➖", "formula": "=H17", "fill": fill_blue, "font": font_blue_kpi},
        {"col": 4, "label": "CRECIENDO 📈", "formula": "=G17", "fill": fill_green, "font": font_green_kpi}
    ]
    
    ws_dash.row_dimensions[4].height = 18
    ws_dash.row_dimensions[5].height = 28
    ws_dash.row_dimensions[6].height = 18
    
    for k in kpis_config:
        c1 = ws_dash.cell(row=4, column=k['col'], value=k['label'])
        c1.font = label_font
        c1.alignment = center_align
        
        c2 = ws_dash.cell(row=5, column=k['col'], value=k['formula'])
        c2.font = k['font']
        c2.alignment = center_align
        
        c3 = ws_dash.cell(row=6, column=k['col'], value=f"={get_column_letter(k['col'])}5/$G$9")
        c3.font = bold_font
        c3.alignment = center_align
        c3.number_format = '0.0%'
        
        for r in range(4, 7):
            cell = ws_dash.cell(row=r, column=k['col'])
            cell.fill = k['fill']
            cell.border = thin_border
            
    headers_resumen = ["Estado de Salud", "Clientes", "% Participación"]
    for col_offset, h in enumerate(headers_resumen):
        cell = ws_dash.cell(row=4, column=6 + col_offset, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    res_states = ["Ha subido", "Ha bajado", "Constante", "Perdido"]
    for idx, state in enumerate(res_states):
        curr_row = 5 + idx
        ws_dash.row_dimensions[curr_row].height = 20
        
        ws_dash.cell(row=curr_row, column=6, value=state).alignment = left_align
        
        # Referencias correctas a los totales en la fila 17
        if state == "Ha subido":
            col_ref = "G"
        elif state == "Ha bajado":
            col_ref = "I"
        elif state == "Constante":
            col_ref = "H"
        else: # Perdido
            col_ref = "J"
            
        ws_dash.cell(row=curr_row, column=7, value=f"={col_ref}17").alignment = center_align
        
        ws_dash.cell(row=curr_row, column=8, value=f"=G{curr_row}/$G$9").alignment = center_align
        ws_dash.cell(row=curr_row, column=8).number_format = '0.0%'
        
        state_fill = fill_green if idx == 0 else (fill_red if idx == 1 else (fill_blue if idx == 2 else fill_grey))
        for col_idx in range(6, 9):
            cell = ws_dash.cell(row=curr_row, column=col_idx)
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = state_fill
            
    ws_dash.row_dimensions[9].height = 20
    ws_dash.cell(row=9, column=6, value="Total TOP 300").font = bold_font
    ws_dash.cell(row=9, column=6).border = thin_border
    ws_dash.cell(row=9, column=6).fill = zebra_fill
    
    ws_dash.cell(row=9, column=7, value="=SUM(G5:G8)").font = bold_font
    ws_dash.cell(row=9, column=7).alignment = center_align
    ws_dash.cell(row=9, column=7).border = thin_border
    ws_dash.cell(row=9, column=7).fill = zebra_fill
    
    ws_dash.cell(row=9, column=8, value="=SUM(H5:H8)").font = bold_font
    ws_dash.cell(row=9, column=8).alignment = center_align
    ws_dash.cell(row=9, column=8).border = thin_border
    ws_dash.cell(row=9, column=8).fill = zebra_fill
    ws_dash.cell(row=9, column=8).number_format = '0.0%'
    
    ws_dash.cell(row=11, column=6, value="MATRIZ DE RIESGO COMERCIAL TOP 300 (Segmentación ABC vs Salud)").font = Font(name=font_family, size=11, bold=True, color="1F4E78")
    
    headers_matrix = ["Segmento ABC", "Ha subido 📈", "Constante ➖", "Ha bajado 📉", "Perdido 🚨", "Total"]
    for col_offset, h in enumerate(headers_matrix):
        cell = ws_dash.cell(row=13, column=6 + col_offset, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws_dash.row_dimensions[13].height = 26
    
    matrix_rows = ["Clase A", "Clase B", "Clase C"]
    last_row_data = 4 + len(df_final)
    
    for idx, segment in enumerate(matrix_rows):
        curr_row = 14 + idx
        ws_dash.row_dimensions[curr_row].height = 20
        
        c_lbl = ws_dash.cell(row=curr_row, column=6, value=f"{segment} (Alto Valor)" if idx==0 else (f"{segment} (Valor Medio)" if idx==1 else f"{segment} (Valor Bajo)"))
        c_lbl.font = bold_font
        c_lbl.alignment = left_align
        c_lbl.border = thin_border
        c_lbl.fill = zebra_fill
        
        c_sub = ws_dash.cell(row=curr_row, column=7, value=f"=COUNTIFS('Cartera Completa'!$D$5:$D${last_row_data}, \"{segment}\", 'Cartera Completa'!$J$5:$J${last_row_data}, \"Ha subido\")")
        c_con = ws_dash.cell(row=curr_row, column=8, value=f"=COUNTIFS('Cartera Completa'!$D$5:$D${last_row_data}, \"{segment}\", 'Cartera Completa'!$J$5:$J${last_row_data}, \"Constante\")")
        c_baj = ws_dash.cell(row=curr_row, column=9, value=f"=COUNTIFS('Cartera Completa'!$D$5:$D${last_row_data}, \"{segment}\", 'Cartera Completa'!$J$5:$J${last_row_data}, \"Ha bajado\")")
        c_per = ws_dash.cell(row=curr_row, column=10, value=f"=COUNTIFS('Cartera Completa'!$D$5:$D${last_row_data}, \"{segment}\", 'Cartera Completa'!$J$5:$J${last_row_data}, \"Perdido\")")
        
        c_row_tot = ws_dash.cell(row=curr_row, column=11, value=f"=SUM(G{curr_row}:J{curr_row})")
        
        for col_idx in range(7, 12):
            cell = ws_dash.cell(row=curr_row, column=col_idx)
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = thin_border
            
        c_sub.fill = fill_green
        c_con.fill = fill_blue
        c_baj.fill = fill_red if idx < 2 else white_fill
        c_per.fill = fill_grey if idx >= 2 else fill_red
        c_row_tot.font = bold_font
        c_row_tot.fill = zebra_fill
        
    ws_dash.row_dimensions[17].height = 20
    c_tot_lbl = ws_dash.cell(row=17, column=6, value="Total")
    c_tot_lbl.font = bold_font
    c_tot_lbl.border = thin_border
    c_tot_lbl.fill = zebra_fill
    
    for c_offset in range(1, 6):
        col_idx = 6 + c_offset
        col_let = get_column_letter(col_idx)
        cell = ws_dash.cell(row=17, column=col_idx, value=f"=SUM({col_let}14:{col_let}16)")
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = zebra_fill
        
    chart = DoughnutChart()
    chart.style = 10
    chart.title = "Salud General de la Cartera TOP 300 (2026)"
    chart.width = 12.5
    chart.height = 9.5
    
    labels_ref = Reference(ws_dash, min_col=6, min_row=5, max_row=8)
    data_ref = Reference(ws_dash, min_col=7, min_row=4, max_row=8)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    ws_dash.add_chart(chart, "M4")
    
    # ----------------- HOJA 2 & 3: ALERTAS ROJAS Y NARANJAS -----------------
    alert_sheets = [
        {
            'name': 'Alertas Rojas (Directores)',
            'df': df_alertas_rojas,
            'desc': 'Clientes Clase A del TOP 300 con caídas de volumen >15% o perdidos en 2026. Requieren acción comercial inmediata de Directores.'
        },
        {
            'name': 'Alertas Naranjas (Vendedores)',
            'df': df_alertas_naranjas,
            'desc': 'Clientes Clase B del TOP 300 con caídas de volumen >15% o perdidos en 2026. Requieren seguimiento prioritario de Vendedores.'
        }
    ]
    
    headers_alerts = [
        'Nombre del Holder', 'Director', 'Vendedor', 'Segmento', 
        'Promedio Inicial', 'Promedio Reciente', 'Variación', 
        'Pérdida Absoluta ($)', 'Health Score', 'Acción Comercial Sugerida'
    ]
    
    for ash in alert_sheets:
        ws = wb.create_sheet(title=ash['name'])
        ws.views.sheetView[0].showGridLines = True
        df = ash['df']
        
        ws['A1'] = ash['name'].upper()
        ws['A1'].font = title_font
        ws['A2'] = ash['desc']
        ws['A2'].font = desc_font
        
        for c_idx, h in enumerate(headers_alerts, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[4].height = 26
        
        for r_idx, r in df.reset_index(drop=True).iterrows():
            curr_row = 5 + r_idx
            ws.row_dimensions[curr_row].height = 20
            row_fill = zebra_fill if r_idx % 2 == 1 else white_fill
            
            ws.cell(row=curr_row, column=1, value=r['Holder']).alignment = left_align
            ws.cell(row=curr_row, column=2, value=r['Director']).alignment = center_align
            ws.cell(row=curr_row, column=3, value=r['Vendedor']).alignment = center_align
            ws.cell(row=curr_row, column=4, value=r['Segmento']).alignment = center_align
            
            c5 = ws.cell(row=curr_row, column=5, value=r['Promedio Inicial'])
            c5.alignment = right_align
            c5.number_format = '#,##0.00'
            
            c6 = ws.cell(row=curr_row, column=6, value=r['Promedio Reciente'])
            c6.alignment = right_align
            c6.number_format = '#,##0.00'
            
            c7 = ws.cell(row=curr_row, column=7, value=r['Variación'] / 100.0)
            c7.alignment = right_align
            c7.number_format = '+0.0%;-0.0%;0.0%'
            c7.fill = fill_red
            
            c8 = ws.cell(row=curr_row, column=8, value=r['Pérdida Absoluta'])
            c8.alignment = right_align
            c8.number_format = '#,##0.00'
            c8.fill = fill_red
            
            c9 = ws.cell(row=curr_row, column=9, value=r['Health Score'])
            c9.alignment = center_align
            c9.number_format = '0'
            c9.font = bold_font
            
            c9.fill = fill_red if r['Health Score'] < 50.0 else fill_grey
            
            ws.cell(row=curr_row, column=10, value=r['Acción Comercial Sugerida']).alignment = left_align
            
            for col_idx in range(1, 11):
                cell = ws.cell(row=curr_row, column=col_idx)
                if col_idx not in [7, 8, 9]:
                    cell.fill = row_fill
                cell.font = normal_font
                cell.border = thin_border
                
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 22
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 75
        
    # ----------------- HOJA 4: CARTERA COMPLETA -----------------
    ws_full = wb.create_sheet(title="Cartera Completa")
    ws_full.views.sheetView[0].showGridLines = True
    
    ws_full['A1'] = "CARTERA COMPLETA TOP 300 2026 - HEALTH SCORE"
    ws_full['A1'].font = title_font
    ws_full['A2'] = f"Listado detallado de los {n_clients} comercios del TOP 300 activos en 2026"
    ws_full['A2'].font = desc_font
    
    headers_full = [
        'Nombre del Holder', 'Director', 'Vendedor', 'Segmento ABC', 
        'Promedio Inicial (W1-21)', 'Promedio Reciente (W22-25)', 'Variación (%)', 
        'Pérdida Absoluta ($)', 'Health Score (0-100)', 'Estado de Salud', 
        'Prioridad de Atención', 'Acción Comercial Sugerida', 'Última Semana Activo', 'Último Mes Activo'
    ]
    
    for c_idx, h in enumerate(headers_full, start=1):
        cell = ws_full.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws_full.row_dimensions[4].height = 26
    
    for r_idx, r in df_final.iterrows():
        curr_row = 5 + r_idx
        ws_full.row_dimensions[curr_row].height = 20
        row_fill = zebra_fill if r_idx % 2 == 1 else white_fill
        
        ws_full.cell(row=curr_row, column=1, value=r['Holder']).alignment = left_align
        ws_full.cell(row=curr_row, column=2, value=r['Director']).alignment = center_align
        ws_full.cell(row=curr_row, column=3, value=r['Vendedor']).alignment = center_align
        ws_full.cell(row=curr_row, column=4, value=r['Segmento']).alignment = center_align
        
        c5 = ws_full.cell(row=curr_row, column=5, value=r['Promedio Inicial'])
        c5.alignment = right_align
        c5.number_format = '#,##0.00'
        
        c6 = ws_full.cell(row=curr_row, column=6, value=r['Promedio Reciente'])
        c6.alignment = right_align
        c6.number_format = '#,##0.00'
        
        c7 = ws_full.cell(row=curr_row, column=7, value=r['Variación'] / 100.0)
        c7.alignment = right_align
        c7.number_format = '+0.0%;-0.0%;0.0%'
        
        c8 = ws_full.cell(row=curr_row, column=8, value=r['Pérdida Absoluta'])
        c8.alignment = right_align
        c8.number_format = '#,##0.00'
        
        c9 = ws_full.cell(row=curr_row, column=9, value=r['Health Score'])
        c9.alignment = center_align
        c9.number_format = '0'
        c9.font = bold_font
        
        ws_full.cell(row=curr_row, column=10, value=r['Estado de Salud']).alignment = center_align
        ws_full.cell(row=curr_row, column=11, value=r['Prioridad de Atención']).alignment = center_align
        ws_full.cell(row=curr_row, column=12, value=r['Acción Comercial Sugerida']).alignment = left_align
        
        ws_full.cell(row=curr_row, column=13, value=r['Última Semana Activo']).alignment = center_align
        ws_full.cell(row=curr_row, column=14, value=r['Último Mes Activo']).alignment = center_align
        
        for col_idx in range(1, 15):
            cell = ws_full.cell(row=curr_row, column=col_idx)
            cell.font = normal_font
            cell.fill = row_fill
            cell.border = thin_border
            
    ws_full.column_dimensions['A'].width = 32
    ws_full.column_dimensions['B'].width = 22
    ws_full.column_dimensions['C'].width = 22
    ws_full.column_dimensions['D'].width = 15
    ws_full.column_dimensions['E'].width = 25
    ws_full.column_dimensions['F'].width = 25
    ws_full.column_dimensions['G'].width = 15
    ws_full.column_dimensions['H'].width = 22
    ws_full.column_dimensions['I'].width = 22
    ws_full.column_dimensions['J'].width = 16
    ws_full.column_dimensions['K'].width = 35
    ws_full.column_dimensions['L'].width = 75
    ws_full.column_dimensions['M'].width = 22
    ws_full.column_dimensions['N'].width = 20
    
    print("Aplicando semáforos condicionales en Cartera Completa...")
    score_rule = ColorScaleRule(start_type='num', start_value=0, start_color='F8CBAD',
                                 mid_type='num', mid_value=50, mid_color='FFF2CC',
                                 end_type='num', end_value=100, end_color='C6E0B4')
    ws_full.conditional_formatting.add(f"I5:I{last_row_data}", score_rule)
    
    var_rule = ColorScaleRule(start_type='num', start_value=-0.50, start_color='F8CBAD',
                              mid_type='num', mid_value=0.0, mid_color='FFFFFF',
                              end_type='num', end_value=0.50, end_color='C6E0B4')
    ws_full.conditional_formatting.add(f"G5:G{last_row_data}", var_rule)
    
    # 6. Guardar archivo
    try:
        wb.save(OUTPUT_FILE)
        print(f"Reporte de Health Score TOP 300 generado con éxito en '{OUTPUT_FILE}'")
    except PermissionError:
        import time
        fallback_name = f"Reporte_Salud_Top300_2026_{int(time.time())}.xlsx"
        try:
            wb.save(fallback_name)
            print(f"Advertencia: '{OUTPUT_FILE}' estaba abierto en otro programa.")
            print(f"Reporte guardado con éxito en su lugar como: '{fallback_name}'")
        except Exception as ex:
            print(f"Error al intentar guardar en fallback '{fallback_name}': {ex}")
            raise

if __name__ == '__main__':
    generar_reporte_top300()
