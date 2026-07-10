import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import DoughnutChart, Reference

# Configuración del script
INPUT_FILE = 'TRX WU_BP.xlsx'
OUTPUT_FILE = 'Reporte_Salud_Clientes_2026.xlsx'
INACTIVITY_WEEKS = 4  # Umbral de semanas inactivas para considerar perdido en 2026
VARIATION_THRESHOLD = 15.0  # Umbral de variación porcentual para clasificar salud

def generar_reporte_salud():
    print("Iniciando análisis de Health Score de la cartera 2026...")
    if not os.path.exists(INPUT_FILE):
        print(f"Error: No se encontró el archivo de entrada '{INPUT_FILE}'")
        return
        
    # 1. Cargar la hoja Semáforo
    print("Cargando hoja de 'Semáforo'...")
    df_sem = pd.read_excel(INPUT_FILE, sheet_name='Semáforo')
    cols = df_sem.columns.tolist()
    row_0 = df_sem.iloc[0].tolist()
    
    # 2. Reconstruir dataframe limpio
    df_sem_clean = df_sem.iloc[1:].copy()
    df_sem_clean.rename(columns={cols[0]: 'Director', cols[1]: 'Holder'}, inplace=True)
    df_sem_clean = df_sem_clean[df_sem_clean['Holder'].notna()]
    df_sem_clean = df_sem_clean[df_sem_clean['Holder'].astype(str).str.lower() != 'total']
    
    # Filtrar semanas de 2026
    weeks_2026 = [c for c in cols if str(c).startswith('Semana') and '2026' in str(c)]
    
    # Convertir semanas de 2026 a numérico y llenar NaNs con 0
    for col in weeks_2026:
        df_sem_clean[col] = pd.to_numeric(df_sem_clean[col], errors='coerce').fillna(0)
        
    # Filtrar solo clientes con actividad en 2026
    df_26 = df_sem_clean[df_sem_clean[weeks_2026].sum(axis=1) > 0].copy()
    
    # Mapeo de semana a mes
    week_to_month = {col: month for col, month in zip(cols, row_0) if str(col).startswith('Semana')}
    
    # Límite de semanas en 2026 para considerarse "perdido"
    # El archivo tiene 25 semanas en 2026. Si INACTIVITY_WEEKS = 4,
    # el cliente debe estar inactivo en las semanas 22, 23, 24, 25.
    # Su última semana activa debe ser <= Semana 25 - 4 = Semana 21.
    last_week_limit_2026 = len(weeks_2026) - INACTIVITY_WEEKS
    last_4_weeks = weeks_2026[-4:]
    
    # Listas por categoría
    constantes = []
    subido = []
    bajado = []
    perdidos = []
    
    for idx, row in df_26.iterrows():
        holder = row['Holder']
        director = row['Director'] if pd.notna(row['Director']) else 'Sin Director'
        
        # Encontrar semanas con ventas > 0
        active_w_26 = [w for w in weeks_2026 if row[w] > 0]
        last_w = active_w_26[-1]
        w_num = int(last_w.split()[1])
        
        # Criterio 1: Perdido (0 transacciones en las últimas 4 semanas de 2026)
        recent_activity = sum(row[last_4_weeks]) > 0
        if not recent_activity:
            avg_historical = float(row[active_w_26].mean())
            perdidos.append({
                'Holder': holder,
                'Director': director,
                'Última Semana Activo': last_w,
                'Último Mes Activo': week_to_month[last_w],
                'Promedio Histórico': avg_historical
            })
        else:
            # Criterio 2: Activo (calcular variación)
            first_w = active_w_26[0]
            first_w_idx = weeks_2026.index(first_w)
            
            if first_w_idx >= 21:
                # Si empezó en la semana 22 o posterior, es muy nuevo
                constantes.append({
                    'Holder': holder,
                    'Director': director,
                    'Promedio Inicial': float(row[first_w]),
                    'Promedio Reciente': float(row[last_4_weeks].mean()),
                    'Variación': 0.0
                })
            else:
                # Baseline: desde su primera semana de 2026 hasta semana 21
                baseline_weeks = weeks_2026[first_w_idx:21]
                avg_baseline = float(row[baseline_weeks].mean())
                avg_recent = float(row[last_4_weeks].mean())
                
                if avg_baseline == 0:
                    var = 0.0
                else:
                    var = (avg_recent - avg_baseline) / avg_baseline * 100
                    
                client_info = {
                    'Holder': holder,
                    'Director': director,
                    'Promedio Inicial': avg_baseline,
                    'Promedio Reciente': avg_recent,
                    'Variación': var
                }
                
                if var > VARIATION_THRESHOLD:
                    subido.append(client_info)
                elif var < -VARIATION_THRESHOLD:
                    bajado.append(client_info)
                else:
                    constantes.append(client_info)
                    
    # Convertir a DataFrames y ordenar
    df_subido = pd.DataFrame(subido).sort_values(by='Variación', ascending=False) if subido else pd.DataFrame(columns=['Holder', 'Director', 'Promedio Inicial', 'Promedio Reciente', 'Variación'])
    df_bajado = pd.DataFrame(bajado).sort_values(by='Variación', ascending=True) if bajado else pd.DataFrame(columns=['Holder', 'Director', 'Promedio Inicial', 'Promedio Reciente', 'Variación'])
    df_constante = pd.DataFrame(constantes).sort_values(by='Promedio Reciente', ascending=False) if constantes else pd.DataFrame(columns=['Holder', 'Director', 'Promedio Inicial', 'Promedio Reciente', 'Variación'])
    df_perdido = pd.DataFrame(perdidos).sort_values(by='Promedio Histórico', ascending=False) if perdidos else pd.DataFrame(columns=['Holder', 'Director', 'Última Semana Activo', 'Último Mes Activo', 'Promedio Histórico'])
    
    print(f"Resultados del Health Score 2026:")
    print(f"  - Ha subido: {len(df_subido)} clientes")
    print(f"  - Ha bajado: {len(df_bajado)} clientes")
    print(f"  - Constante: {len(df_constante)} clientes")
    print(f"  - Perdido: {len(df_perdido)} clientes")
    
    # 4. Crear archivo Excel Premium
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="1F4E78")
    desc_font = Font(name=font_family, size=9, italic=True, color="595959")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    zebra_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Colores para las tarjetas KPI y estados
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    font_green = Font(name=font_family, size=18, bold=True, color="375623")
    font_red = Font(name=font_family, size=18, bold=True, color="C65911")
    font_blue = Font(name=font_family, size=18, bold=True, color="1F4E78")
    font_grey = Font(name=font_family, size=18, bold=True, color="595959")
    
    label_font = Font(name=font_family, size=9, bold=True, color="595959")
    normal_font = Font(name=font_family, size=10)
    bold_font = Font(name=font_family, size=10, bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # ----------------- PESTAÑA 1: DASHBOARD RESUMEN -----------------
    ws_dash = wb.create_sheet(title="Resumen de Salud")
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash['A1'] = "HEALTH SCORE DE LA CARTERA 2026"
    ws_dash['A1'].font = title_font
    ws_dash['A2'] = f"Clasificación de salud de los {len(df_26)} clientes con actividad durante el 2026 (Umbral de variación: {VARIATION_THRESHOLD}%)"
    ws_dash['A2'].font = desc_font
    
    # Diseñar tarjetas KPI superiores (A4:E6)
    kpis = [
        {"col": 1, "label": "CRECIENDO 📈", "formula": "=H5", "pct_formula": "=I5", "fill": fill_green, "font": font_green},
        {"col": 2, "label": "EN CAÍDA 📉", "formula": "=H6", "pct_formula": "=I6", "fill": fill_red, "font": font_red},
        {"col": 3, "label": "ESTABLES ➖", "formula": "=H7", "pct_formula": "=I7", "fill": fill_blue, "font": font_blue},
        {"col": 4, "label": "BAJAS 🚨", "formula": "=H8", "pct_formula": "=I8", "fill": fill_grey, "font": font_grey},
        {"col": 5, "label": "CARTERA SANA 💚", "formula": "=H5+H7", "pct_formula": "=(H5+H7)/H9", "fill": fill_green, "font": font_green}
    ]
    
    ws_dash.row_dimensions[4].height = 18
    ws_dash.row_dimensions[5].height = 28
    ws_dash.row_dimensions[6].height = 18
    
    for k in kpis:
        # Fila 4: Label
        c1 = ws_dash.cell(row=4, column=k['col'], value=k['label'])
        c1.font = label_font
        c1.alignment = center_align
        
        # Fila 5: Valor absoluto
        c2 = ws_dash.cell(row=5, column=k['col'], value=k['formula'])
        c2.font = k['font']
        c2.alignment = center_align
        
        # Fila 6: Porcentaje
        c3 = ws_dash.cell(row=6, column=k['col'], value=k['pct_formula'])
        c3.font = bold_font
        c3.alignment = center_align
        c3.number_format = '0.0%'
        
        # Formatos comunes para la tarjeta
        for r_idx in range(4, 7):
            cell = ws_dash.cell(row=r_idx, column=k['col'])
            cell.fill = k['fill']
            cell.border = thin_border
            
    # Escribir Tabla de Resumen (G4:I9)
    headers_resumen = ["Estado de Salud", "Clientes", "% Participación"]
    for col_offset, h in enumerate(headers_resumen):
        cell = ws_dash.cell(row=4, column=7 + col_offset, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    resumen_datos = [
        ("Ha subido", len(df_subido)),
        ("Ha bajado", len(df_bajado)),
        ("Constante", len(df_constante)),
        ("Perdido", len(df_perdido))
    ]
    
    for idx, (estado, cant) in enumerate(resumen_datos):
        curr_row = 5 + idx
        ws_dash.row_dimensions[curr_row].height = 20
        
        # Estado
        c_est = ws_dash.cell(row=curr_row, column=7, value=estado)
        c_est.alignment = left_align
        
        # Cantidad
        c_cant = ws_dash.cell(row=curr_row, column=8, value=cant)
        c_cant.alignment = center_align
        
        # %
        c_pct = ws_dash.cell(row=curr_row, column=9, value=f"=H{curr_row}/$H$9")
        c_pct.alignment = center_align
        c_pct.number_format = '0.0%'
        
        # Colorear según el estado para coherencia visual
        state_fill = fill_green if idx == 0 else (fill_red if idx == 1 else (fill_blue if idx == 2 else fill_grey))
        for col_idx in range(7, 10):
            cell = ws_dash.cell(row=curr_row, column=col_idx)
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = state_fill
            
    # Fila de Total Resumen
    total_row = 9
    ws_dash.row_dimensions[total_row].height = 20
    
    c_tot_lbl = ws_dash.cell(row=total_row, column=7, value="Total Cartera")
    c_tot_lbl.font = bold_font
    c_tot_lbl.alignment = left_align
    c_tot_lbl.border = thin_border
    c_tot_lbl.fill = zebra_fill
    
    c_tot_val = ws_dash.cell(row=total_row, column=8, value="=SUM(H5:H8)")
    c_tot_val.font = bold_font
    c_tot_val.alignment = center_align
    c_tot_val.border = thin_border
    c_tot_val.fill = zebra_fill
    
    c_tot_pct = ws_dash.cell(row=total_row, column=9, value="=SUM(I5:I8)")
    c_tot_pct.font = bold_font
    c_tot_pct.alignment = center_align
    c_tot_pct.border = thin_border
    c_tot_pct.fill = zebra_fill
    c_tot_pct.number_format = '0.0%'
    
    # Crear Gráfico de Doughnut (K4)
    chart = DoughnutChart()
    chart.style = 10
    chart.title = "Distribución de Salud de la Cartera 2026"
    chart.width = 13
    chart.height = 9.5
    
    labels_ref = Reference(ws_dash, min_col=7, min_row=5, max_row=8)
    data_ref = Reference(ws_dash, min_col=8, min_row=4, max_row=8)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    
    # Añadir al dashboard
    ws_dash.add_chart(chart, "K4")
    
    # Ajustar anchos del dashboard
    ws_dash.column_dimensions['A'].width = 16
    ws_dash.column_dimensions['B'].width = 16
    ws_dash.column_dimensions['C'].width = 16
    ws_dash.column_dimensions['D'].width = 16
    ws_dash.column_dimensions['E'].width = 18
    ws_dash.column_dimensions['F'].width = 4
    ws_dash.column_dimensions['G'].width = 18
    ws_dash.column_dimensions['H'].width = 12
    ws_dash.column_dimensions['I'].width = 16
    ws_dash.column_dimensions['J'].width = 4
    
    # ----------------- PESTAÑAS 2-4: HA SUBIDO, HA BAJADO, CONSTANTE -----------------
    active_sheets_config = [
        {"name": "Ha Subido", "df": df_subido, "desc": f"Clientes con transacciones en 2026 que crecieron más de {VARIATION_THRESHOLD}% frente a su promedio histórico", "fill": fill_green},
        {"name": "Ha Bajado", "df": df_bajado, "desc": f"Clientes con transacciones en 2026 que cayeron más de {VARIATION_THRESHOLD}% frente a su promedio histórico", "fill": fill_red},
        {"name": "Constante", "df": df_constante, "desc": f"Clientes con transacciones en 2026 y volumen de operaciones estable (dentro de ±{VARIATION_THRESHOLD}%)", "fill": fill_blue}
    ]
    
    headers_active = ['Nombre del Holder', 'Director', 'Promedio Inicial (W1-21)', 'Promedio Reciente (W22-25)', 'Variación (%)']
    
    for sh in active_sheets_config:
        ws = wb.create_sheet(title=sh['name'])
        ws.views.sheetView[0].showGridLines = True
        df = sh['df']
        
        # Títulos
        ws['A1'] = sh['name'].upper()
        ws['A1'].font = title_font
        ws['A2'] = sh['desc']
        ws['A2'].font = desc_font
        
        # Headers tabla principal
        for c_idx, h in enumerate(headers_active, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[4].height = 26
        
        # Escribir datos
        for r_idx, r in df.reset_index(drop=True).iterrows():
            curr_row = 5 + r_idx
            ws.row_dimensions[curr_row].height = 20
            
            row_fill = zebra_fill if r_idx % 2 == 1 else white_fill
            
            # Holder
            c1 = ws.cell(row=curr_row, column=1, value=r['Holder'])
            c1.alignment = left_align
            # Director
            c2 = ws.cell(row=curr_row, column=2, value=r['Director'])
            c2.alignment = center_align
            # Promedio Inicial
            c3 = ws.cell(row=curr_row, column=3, value=r['Promedio Inicial'])
            c3.alignment = right_align
            c3.number_format = '#,##0.00'
            # Promedio Reciente
            c4 = ws.cell(row=curr_row, column=4, value=r['Promedio Reciente'])
            c4.alignment = right_align
            c4.number_format = '#,##0.00'
            # Variación (Fórmula en porcentaje directo)
            # En openpyxl escribimos la variación dividida entre 100 para formatear como porcentaje excel
            c5 = ws.cell(row=curr_row, column=5, value=r['Variación'] / 100.0)
            c5.alignment = right_align
            c5.number_format = '+0.0%;-0.0%;0.0%'  # Fuerza el signo + para incrementos y - para decrementos
            
            # Estilos comunes de la celda
            for col_idx in range(1, 6):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = normal_font
                cell.fill = row_fill
                cell.border = thin_border
                
            # Colorear la celda de Variación con el color de salud correspondiente
            c5.fill = sh['fill']
            
        # Formatear anchos de columnas
        ws.column_dimensions['A'].width = 35 # Holder
        ws.column_dimensions['B'].width = 25 # Director
        ws.column_dimensions['C'].width = 25 # Promedio Inicial
        ws.column_dimensions['D'].width = 25 # Promedio Reciente
        ws.column_dimensions['E'].width = 18 # Variación
        
    # ----------------- PESTAÑA 5: PERDIDOS -----------------
    ws_perd = wb.create_sheet(title="Perdidos")
    ws_perd.views.sheetView[0].showGridLines = True
    
    ws_perd['A1'] = "CLIENTES PERDIDOS EN 2026"
    ws_perd['A2'] = f"Clientes activos en 2026 que registran 0 actividad en las últimas 4 semanas (Semanas 22 a 25)"
    ws_perd['A1'].font = title_font
    ws_perd['A2'].font = desc_font
    
    headers_perd = ['Nombre del Holder', 'Director', 'Última Semana Activo', 'Último Mes Activo', 'Promedio Semanal Histórico']
    for c_idx, h in enumerate(headers_perd, start=1):
        cell = ws_perd.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws_perd.row_dimensions[4].height = 26
    
    for r_idx, r in df_perdido.reset_index(drop=True).iterrows():
        curr_row = 5 + r_idx
        ws_perd.row_dimensions[curr_row].height = 20
        row_fill = zebra_fill if r_idx % 2 == 1 else white_fill
        
        # Holder
        c1 = ws_perd.cell(row=curr_row, column=1, value=r['Holder'])
        c1.alignment = left_align
        # Director
        c2 = ws_perd.cell(row=curr_row, column=2, value=r['Director'])
        c2.alignment = center_align
        # Última Semana
        c3 = ws_perd.cell(row=curr_row, column=3, value=r['Última Semana Activo'])
        c3.alignment = center_align
        # Último Mes
        c4 = ws_perd.cell(row=curr_row, column=4, value=r['Último Mes Activo'])
        c4.alignment = center_align
        # Promedio Histórico
        c5 = ws_perd.cell(row=curr_row, column=5, value=r['Promedio Histórico'])
        c5.alignment = right_align
        c5.number_format = '#,##0.00'
        
        # Estilos comunes de la celda
        for col_idx in range(1, 6):
            cell = ws_perd.cell(row=curr_row, column=col_idx)
            cell.font = normal_font
            cell.fill = row_fill
            cell.border = thin_border
            
        # Resaltar la última semana y mes en gris como inactivo
        c3.fill = fill_grey
        c4.fill = fill_grey
        
    ws_perd.column_dimensions['A'].width = 35 # Holder
    ws_perd.column_dimensions['B'].width = 25 # Director
    ws_perd.column_dimensions['C'].width = 22 # Última semana
    ws_perd.column_dimensions['D'].width = 20 # Último mes
    ws_perd.column_dimensions['E'].width = 25 # Promedio Histórico
    
    # 5. Guardar archivo con control de error por bloqueo
    try:
        wb.save(OUTPUT_FILE)
        print(f"Reporte de Health Score generado con éxito en '{OUTPUT_FILE}'")
    except PermissionError:
        import time
        fallback_name = f"Reporte_Salud_Clientes_2026_{int(time.time())}.xlsx"
        try:
            wb.save(fallback_name)
            print(f"Advertencia: '{OUTPUT_FILE}' estaba abierto en otro programa.")
            print(f"Reporte guardado con éxito en su lugar como: '{fallback_name}'")
        except Exception as ex:
            print(f"Error al intentar guardar en fallback '{fallback_name}': {ex}")
            raise

if __name__ == '__main__':
    generar_reporte_salud()
