import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Configuración del script
INPUT_FILE = 'TRX WU_BP.xlsx'
OUTPUT_FILE = 'Reporte_Clientes_Perdidos.xlsx'
INACTIVITY_WEEKS = 4  # Umbral de semanas inactivas en 2026

def generar_reporte():
    print("Iniciando análisis de clientes perdidos...")
    if not os.path.exists(INPUT_FILE):
        print(f"Error: No se encontró el archivo de entrada '{INPUT_FILE}'")
        return
        
    # 1. Cargar la hoja Semáforo
    print("Cargando hoja de 'Semáforo'...")
    df_sem = pd.read_excel(INPUT_FILE, sheet_name='Semáforo')
    cols = df_sem.columns.tolist()
    row_0 = df_sem.iloc[0].tolist()
    
    # 2. Reconstruir dataframe limpio
    df_sem_clean = df_sem.iloc[1:].copy()
    df_sem_clean.rename(columns={cols[0]: 'Director', cols[1]: 'Holder'}, inplace=True)
    
    # Filtrar filas vacías o totales
    df_sem_clean = df_sem_clean[df_sem_clean['Holder'].notna()]
    df_sem_clean = df_sem_clean[df_sem_clean['Holder'].astype(str).str.lower() != 'total']
    
    # Identificar las semanas y mapear a meses
    weeks_2025 = [c for c in cols if str(c).startswith('Semana') and '2026' not in str(c)]
    weeks_2026 = [c for c in cols if str(c).startswith('Semana') and '2026' in str(c)]
    
    # Convertir a numérico y llenar NaNs con 0
    all_weeks = weeks_2025 + weeks_2026
    for col in all_weeks:
        df_sem_clean[col] = pd.to_numeric(df_sem_clean[col], errors='coerce').fillna(0)
        
    # Mapeo de semana a mes
    week_to_month = {col: month for col, month in zip(cols, row_0) if str(col).startswith('Semana')}
    
    # Calcular clientes activos por mes en toda la base de datos
    # 2025
    weeks_by_month_2025 = {}
    for w in weeks_2025:
        m = week_to_month[w]
        weeks_by_month_2025.setdefault(m, []).append(w)
        
    active_by_month_2025 = {}
    for m, w_list in weeks_by_month_2025.items():
        active_by_month_2025[m] = int((df_sem_clean[w_list].sum(axis=1) > 0).sum())
        
    # 2026
    weeks_by_month_2026 = {}
    for w in weeks_2026:
        m = week_to_month[w]
        weeks_by_month_2026.setdefault(m, []).append(w)
        
    active_by_month_2026 = {}
    for m, w_list in weeks_by_month_2026.items():
        active_by_month_2026[m] = int((df_sem_clean[w_list].sum(axis=1) > 0).sum())
        
    # Calcular Cartera de Clientes (Portfolio) por mes
    # Un cliente está en cartera en el mes M si su primera transacción fue antes o durante el mes M
    # y su última transacción fue en el mes M o posterior.
    week_pos = {w: idx for idx, w in enumerate(all_weeks)}
    
    client_lifetimes = []
    for idx, row in df_sem_clean.iterrows():
        active_w = [w for w in all_weeks if row[w] > 0]
        if active_w:
            client_lifetimes.append({
                'first_pos': week_pos[active_w[0]],
                'last_pos': week_pos[active_w[-1]]
            })
            
    # Portfolio para meses 2025
    portfolio_by_month_2025 = {}
    for m in ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']:
        m_weeks = weeks_by_month_2025.get(m, [])
        if not m_weeks:
            portfolio_by_month_2025[m] = 0
            continue
        first_w_pos = week_pos[m_weeks[0]]
        last_w_pos = week_pos[m_weeks[-1]]
        
        count = 0
        for c in client_lifetimes:
            if c['first_pos'] <= last_w_pos and c['last_pos'] >= first_w_pos:
                count += 1
        portfolio_by_month_2025[m] = count
        
    # Portfolio para meses 2026
    portfolio_by_month_2026 = {}
    for m in ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio']:
        m_weeks = weeks_by_month_2026.get(m, [])
        if not m_weeks:
            portfolio_by_month_2026[m] = 0
            continue
        first_w_pos = week_pos[m_weeks[0]]
        last_w_pos = week_pos[m_weeks[-1]]
        
        count = 0
        for c in client_lifetimes:
            if c['first_pos'] <= last_w_pos and c['last_pos'] >= first_w_pos:
                count += 1
        portfolio_by_month_2026[m] = count
        
    # Calcular totales únicos de activos
    total_active_unique_2025 = int((df_sem_clean[weeks_2025].sum(axis=1) > 0).sum())
    total_active_unique_2026 = int((df_sem_clean[weeks_2026].sum(axis=1) > 0).sum())
    
    # 3. Clasificación de clientes
    lost_2025_data = []
    active_2025_lost_2026_data = []
    active_2026_lost_2026_data = []
    
    # Límite de semanas en 2026 para considerarse "perdido"
    last_week_limit_2026 = len(weeks_2026) - INACTIVITY_WEEKS
    
    for idx, row in df_sem_clean.iterrows():
        holder = row['Holder']
        director = row['Director'] if pd.notna(row['Director']) else 'Sin Director'
        
        # Encontrar semanas con ventas > 0
        active_weeks_25 = [w for w in weeks_2025 if row[w] > 0]
        active_weeks_26 = [w for w in weeks_2026 if row[w] > 0]
        
        has_2025 = len(active_weeks_25) > 0
        has_2026 = len(active_weeks_26) > 0
        
        if has_2025 and not has_2026:
            # Categoría 1: Clientes Perdidos en 2025
            last_w = active_weeks_25[-1]
            last_m = week_to_month[last_w]
            lost_2025_data.append({
                'Holder': holder,
                'Director': director,
                'Última Semana Activo': last_w,
                'Último Mes Activo': last_m
            })
            
        elif has_2026:
            last_w = active_weeks_26[-1]
            w_num = int(last_w.split()[1])
            
            if w_num <= last_week_limit_2026:
                last_m = week_to_month[last_w]
                client_info = {
                    'Holder': holder,
                    'Director': director,
                    'Última Semana Activo': last_w,
                    'Último Mes Activo': last_m
                }
                
                if has_2025:
                    # Categoría 2: Activo en 2025 y se perdió en 2026
                    active_2025_lost_2026_data.append(client_info)
                else:
                    # Categoría 3: Activo solo en 2026 y se perdió en 2026
                    active_2026_lost_2026_data.append(client_info)
                    
    # Convertir a DataFrames
    df_cat1 = pd.DataFrame(lost_2025_data)
    df_cat2 = pd.DataFrame(active_2025_lost_2026_data)
    df_cat3 = pd.DataFrame(active_2026_lost_2026_data)
    
    # Asegurar que no estén vacíos
    for df_tmp in [df_cat1, df_cat2, df_cat3]:
        if df_tmp.empty:
            df_tmp['Holder'] = []
            df_tmp['Director'] = []
            df_tmp['Última Semana Activo'] = []
            df_tmp['Último Mes Activo'] = []
            
    print(f"Clasificación completada:")
    print(f"  - Perdidos 2025: {len(df_cat1)} clientes")
    print(f"  - Activos 2025, Perdidos 2026: {len(df_cat2)} clientes")
    print(f"  - Activos 2026, Perdidos 2026: {len(df_cat3)} clientes")
    
    # 4. Crear archivo Excel Premium
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Hojas y configuraciones
    hojas_config = [
        {
            'name': 'Perdidos 2025',
            'df': df_cat1,
            'desc': 'Clientes activos en 2025 que no han registrado transacciones en todo el 2026',
            'months_order': ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'],
            'active_by_month': active_by_month_2025,
            'portfolio_by_month': portfolio_by_month_2025,
            'total_active_unique': total_active_unique_2025
        },
        {
            'name': 'Activos 2025 - Perdidos 2026',
            'df': df_cat2,
            'desc': f'Clientes activos en 2025 y 2026 que llevan al menos {INACTIVITY_WEEKS} semanas inactivos',
            'months_order': ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio'],
            'active_by_month': active_by_month_2026,
            'portfolio_by_month': portfolio_by_month_2026,
            'total_active_unique': total_active_unique_2026
        },
        {
            'name': 'Activos 2026 - Perdidos 2026',
            'df': df_cat3,
            'desc': f'Clientes nuevos en 2026 que llevan al menos {INACTIVITY_WEEKS} semanas inactivos',
            'months_order': ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio'],
            'active_by_month': active_by_month_2026,
            'portfolio_by_month': portfolio_by_month_2026,
            'total_active_unique': total_active_unique_2026
        }
    ]
    
    # Estilos
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="1F4E78")
    desc_font = Font(name=font_family, size=9, italic=True, color="595959")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    normal_font = Font(name=font_family, size=10)
    bold_font = Font(name=font_family, size=10, bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    for config in hojas_config:
        ws = wb.create_sheet(title=config['name'])
        ws.views.sheetView[0].showGridLines = True
        df = config['df']
        
        # A) Título y Descripción
        ws['A1'] = config['name'].upper()
        ws['A1'].font = title_font
        
        ws['A2'] = config['desc']
        ws['A2'].font = desc_font
        
        # B) Escribir Tabla Principal (A4:D...)
        headers_main = ['Nombre del Holder', 'Director', 'Última Semana Activo', 'Último Mes Activo']
        for col_idx, h in enumerate(headers_main, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        ws.row_dimensions[4].height = 26
        
        # Escribir filas de datos
        data_start_row = 5
        for row_idx, r in df.iterrows():
            curr_row = data_start_row + row_idx
            ws.row_dimensions[curr_row].height = 20
            
            fill_to_use = zebra_fill if row_idx % 2 == 1 else white_fill
            
            # Holder
            c1 = ws.cell(row=curr_row, column=1, value=r['Holder'])
            c1.alignment = left_align
            # Director
            c2 = ws.cell(row=curr_row, column=2, value=r['Director'])
            c2.alignment = center_align
            # Última Semana
            c3 = ws.cell(row=curr_row, column=3, value=r['Última Semana Activo'])
            c3.alignment = center_align
            # Último Mes
            c4 = ws.cell(row=curr_row, column=4, value=r['Último Mes Activo'])
            c4.alignment = center_align
            
            # Estilos comunes
            for col_idx in range(1, 5):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = normal_font
                cell.fill = fill_to_use
                cell.border = thin_border
                
        # C) Escribir Tabla de Resumen Mensual (F4:K...)
        # Headers resumen
        headers_resumen = ["Mes", "Total Clientes", "Clientes Activos", "% Activos", "Clientes Perdidos", "% Perdidos"]
        for col_offset, h in enumerate(headers_resumen):
            col_idx = 6 + col_offset  # Columnas F, G, H, I, J, K
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # Calcular pérdidas mensuales
        lost_counts = df['Último Mes Activo'].value_counts()
        
        total_row = 5 + len(config['months_order'])
        
        for s_idx, m in enumerate(config['months_order']):
            curr_row = 5 + s_idx
            
            # F: Mes
            cf = ws.cell(row=curr_row, column=6, value=m)
            cf.alignment = center_align
            
            # G: Total Clientes (Portfolio/Cartera en este mes)
            cg = ws.cell(row=curr_row, column=7, value=config['portfolio_by_month'].get(m, 0))
            cg.alignment = center_align
            
            # H: Clientes Activos (Activos reales en este mes)
            ch = ws.cell(row=curr_row, column=8, value=config['active_by_month'].get(m, 0))
            ch.alignment = center_align
            
            # I: % Activos (Fórmula: Clientes Activos / Total Clientes de este mes)
            # Ejemplo: =H5/G5
            ci = ws.cell(row=curr_row, column=9, value=f"=H{curr_row}/G{curr_row}")
            ci.alignment = center_align
            ci.number_format = '0.0%'
            
            # J: Clientes Perdidos (Valor directo)
            cj = ws.cell(row=curr_row, column=10, value=lost_counts.get(m, 0))
            cj.alignment = center_align
            
            # K: % Perdidos (Fórmula: Clientes Perdidos / Total Clientes de este mes - Churn rate)
            # Ejemplo: =J5/G5
            ck = ws.cell(row=curr_row, column=11, value=f"=J{curr_row}/G{curr_row}")
            ck.alignment = center_align
            ck.number_format = '0.0%'
            
            # Estilos comunes para filas de datos
            for col_idx in range(6, 12):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = normal_font
                cell.border = thin_border
                
        # Total Fila Resumen (F{total_row}:K{total_row})
        ws.row_dimensions[total_row].height = 20
        
        # F: Total label
        c_tot_label = ws.cell(row=total_row, column=6, value="Total Únicos")
        c_tot_label.font = bold_font
        c_tot_label.alignment = center_align
        c_tot_label.border = thin_border
        c_tot_label.fill = zebra_fill
        
        # G: Clientes Únicos Activos en el año
        c_tot_port = ws.cell(row=total_row, column=7, value=config['total_active_unique'])
        c_tot_port.font = bold_font
        c_tot_port.alignment = center_align
        c_tot_port.border = thin_border
        c_tot_port.fill = zebra_fill
        
        # H: Clientes Únicos Activos (mismo que G)
        c_tot_act = ws.cell(row=total_row, column=8, value=config['total_active_unique'])
        c_tot_act.font = bold_font
        c_tot_act.alignment = center_align
        c_tot_act.border = thin_border
        c_tot_act.fill = zebra_fill
        
        # I: % Activos Total (vacío)
        c_tot_pct_act = ws.cell(row=total_row, column=9, value="")
        c_tot_pct_act.border = thin_border
        c_tot_pct_act.fill = zebra_fill
        
        # J: Clientes Perdidos Total (Fórmula de suma)
        c_tot_lost = ws.cell(row=total_row, column=10, value=f"=SUM(J5:J{total_row-1})")
        c_tot_lost.font = bold_font
        c_tot_lost.alignment = center_align
        c_tot_lost.border = thin_border
        c_tot_lost.fill = zebra_fill
        
        # K: % Perdidos Total (Fórmula: Total Perdidos / Total Únicos del Año)
        # Ejemplo: =J17/G17 (representa la tasa de pérdida acumulada del periodo)
        c_tot_pct_lost = ws.cell(row=total_row, column=11, value=f"=J{total_row}/G{total_row}")
        c_tot_pct_lost.font = bold_font
        c_tot_pct_lost.alignment = center_align
        c_tot_pct_lost.border = thin_border
        c_tot_pct_lost.fill = zebra_fill
        c_tot_pct_lost.number_format = '0.0%'
        
        # D) Crear Gráfico de Barras (ahora referenciado a columna J, y posicionado en M4)
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"Distribución de Pérdidas por Mes - {config['name']}"
        chart.y_axis.title = "Cantidad de Clientes"
        chart.x_axis.title = "Mes"
        chart.width = 15
        chart.height = 10
        chart.legend = None
        
        # Referencias: Datos en col J (10), Categorías en col F (6)
        data_ref = Reference(ws, min_col=10, min_row=4, max_row=total_row-1)
        cats_ref = Reference(ws, min_col=6, min_row=5, max_row=total_row-1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        if len(chart.series) > 0:
            chart.series[0].graphicalProperties.solidFill = "1F4E78"
            
        ws.add_chart(chart, "M4")
        
        # E) Autoajustar ancho de columnas
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col[2:]:
                val = str(cell.value or '')
                if col_letter in ['M', 'N', 'O', 'P', 'Q']:
                    continue  # evitar columnas del gráfico
                if len(val) > max_len:
                    max_len = len(val)
            
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Ajustes de ancho manuales finos
        ws.column_dimensions['A'].width = 35 # Holder
        ws.column_dimensions['B'].width = 25 # Director
        ws.column_dimensions['C'].width = 22 # Última semana
        ws.column_dimensions['D'].width = 20 # Último mes
        ws.column_dimensions['E'].width = 4  # Separador
        ws.column_dimensions['F'].width = 15 # Mes
        ws.column_dimensions['G'].width = 15 # Total Clientes
        ws.column_dimensions['H'].width = 16 # Clientes Activos
        ws.column_dimensions['I'].width = 14 # % Activos
        ws.column_dimensions['J'].width = 18 # Clientes Perdidos
        ws.column_dimensions['K'].width = 14 # % Perdidos
        ws.column_dimensions['L'].width = 4  # Separador
        
    try:
        wb.save(OUTPUT_FILE)
        print(f"Reporte generado con éxito en '{OUTPUT_FILE}'")
    except PermissionError:
        import time
        fallback_name = f"Reporte_Clientes_Perdidos_{int(time.time())}.xlsx"
        try:
            wb.save(fallback_name)
            print(f"Advertencia: '{OUTPUT_FILE}' estaba abierto en otro programa.")
            print(f"Reporte guardado con éxito en su lugar como: '{fallback_name}'")
        except Exception as ex:
            print(f"Error al intentar guardar en fallback '{fallback_name}': {ex}")
            raise

if __name__ == '__main__':
    generar_reporte()
